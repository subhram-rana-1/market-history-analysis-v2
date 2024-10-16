import json
from datetime import datetime, date
import requests
from kiteconnect import KiteConnect
from typing import List
from openpyxl import load_workbook

API_KEY = 'p7qy2u03ev8e45pm'
API_SECRETE = '4pamm45xsirewovl8smza5t1qvft0t92'
ACCESS_TOKEN = None
upstox_ts_format = "%Y-%m-%dT%H:%M:%S%z"
date_str_format = "%Y-%m-%d"
time_str_format = "%H:%M:%S"
datetime_str_format = "%Y-%m-%d %H:%M:%S"


def new_kite_connect_client() -> KiteConnect:
    kc: KiteConnect = KiteConnect(
        api_key=API_KEY,
    )

    print("Please login with here and fetch the 'request_token' from redirected "
          "url after successful login : ", kc.login_url())

    request_token: str = input("enter 'request_token': ")

    session_data: dict = kc.generate_session(
        request_token=request_token,
        api_secret=API_SECRETE,
    )

    ACCESS_TOKEN = session_data['access_token']
    kc.set_access_token(ACCESS_TOKEN)

    print('\nkite connect client creation successful !!! ')

    return kc


# output: {"10-20": [3, cum_occurrence, cum_%age], "20-30": [4, cum_occurrence, cum_%age]}
def get_distributions(
        nums: List[float],
        grouping_range_width: int,
) -> dict:
    nums.sort()
    momentum_distributions = {}

    start_range = 0
    end_range = grouping_range_width

    i = 0
    sum_of_occurrence = 0
    tot_cumulative_occurrence = 0

    while i < len(nums):
        if nums[i] <= end_range:
            sum_of_occurrence += 1
        else:
            key = f'{start_range}-{end_range}'
            momentum_distributions[key] = [sum_of_occurrence]
            tot_cumulative_occurrence += sum_of_occurrence

            start_range = end_range
            end_range = start_range + grouping_range_width

            sum_of_occurrence = 0
            i -= 1

        i += 1

    if sum_of_occurrence != 0:
        momentum_distributions[f'{start_range}-{end_range}'] = [sum_of_occurrence]
        tot_cumulative_occurrence += sum_of_occurrence

    # calculate cumulative percentage
    start_range = 0
    end_range = grouping_range_width
    tot_sum_so_far = 0
    while True:
        my_range = f'{start_range}-{end_range}'
        if my_range not in momentum_distributions:
            break

        count = momentum_distributions[my_range][0]
        tot_sum_so_far += count

        cur_percentage = round(tot_sum_so_far / tot_cumulative_occurrence * 100)
        momentum_distributions[my_range].append(tot_sum_so_far)
        momentum_distributions[my_range].append(cur_percentage)

        start_range = end_range
        end_range = start_range + grouping_range_width

    return momentum_distributions


def _clean_worksheet(ws, start_row: int, start_col: int):
    for row in range(start_row, 1000):  # Adjust range to include row 1000
        for col in range(start_col, start_col+4):  # Columns are typically 1-indexed
            cell = ws.cell(row=row, column=col)
            cell.value = None


def write_distribution_result_to_worksheet(
        distribution_data: dict,
        abs_file_path: str,
        sheet_name: str,
        start_row: int,
        start_col: int,
):
    wb = load_workbook(abs_file_path)
    ws = wb[sheet_name]

    _clean_worksheet(ws, start_row, start_col)

    cur_row = start_row
    for group_range, arr in distribution_data.items():
        count = arr[0]
        cum_sum = arr[1]
        cum_percentage = arr[2]

        ws.cell(row=cur_row, column=start_col, value=group_range)
        ws.cell(row=cur_row, column=start_col+1, value=count)
        ws.cell(row=cur_row, column=start_col+2, value=cum_sum)
        ws.cell(row=cur_row, column=start_col+3, value=cum_percentage)

        cur_row += 1

    wb.save(abs_file_path)
    wb.close()


class Candle:
    def __init__(
            self,
            ts: datetime,
            open: float,
            hi: float,
            lo: float,
            close: float,
    ):
        self.ts = ts
        self.open = open
        self.hi = hi
        self.lo = lo
        self.close = close

    @property
    def date(self) -> str:
        return self.ts.date().strftime(date_str_format)

    @property
    def time(self) -> str:
        return self.ts.time().strftime(time_str_format)

    @classmethod
    def from_api_resp_candle_dict(cls, candle_dict: dict):
        return Candle(
            datetime.strptime(candle_dict[0], upstox_ts_format),
            candle_dict[1],
            candle_dict[2],
            candle_dict[3],
            candle_dict[4],
        )

    @property
    def up_move(self) -> float:
        return self.hi - self.open

    @property
    def down_move(self) -> float:
        return self.open - self.lo

    @property
    def body_length(self) -> float:
        return abs(self.open - self.close)

    @property
    def upper_wick_length(self) -> float:
        return abs(self.hi - max(self.open, self.close))

    @property
    def lower_wick_length(self) -> float:
        return abs(self.lo - min(self.open, self.close))


class UpstoxCandlesticksData:
    def __init__(self, candles: List[Candle]):
        self.candles: List[Candle] = candles

    def get_moves(self, direction: str) -> List[float]:
        if direction == 'up':
            return [candle.up_move for candle in self.candles]
        if direction == 'down':
            return [candle.down_move for candle in self.candles]

    def get_candle_body_lengths(self) -> List[float]:
        return [candle.body_length for candle in self.candles]

    def get_candle_upper_wick_lengths(self) -> List[float]:
        return [candle.upper_wick_length for candle in self.candles]

    def get_candle_lower_wick_lengths(self) -> List[float]:
        return [candle.lower_wick_length for candle in self.candles]


class UpstoxCandlestickResponse:
    def __init__(self, upstox_candlesticks_data: UpstoxCandlesticksData):
        self.status = 'success'
        self.data = upstox_candlesticks_data

    @classmethod
    def from_upstox_api_response(cls, resp: dict):
        candles = resp['data']['candles']
        candles = reversed(candles)
        return UpstoxCandlestickResponse(
            upstox_candlesticks_data=UpstoxCandlesticksData(
                [Candle.from_api_resp_candle_dict(candle_dict) for candle_dict in candles]
            ),
        )

    def distribution_for_moves(self, direction: str, bucket_length: int,
                               max_candle_size_for_analysis: float) -> dict:
        moves = self.data.get_moves(direction)

        # IMPORTANT - remove too big candles, they are outliers to this analysis
        moves = [x for x in moves if x <= max_candle_size_for_analysis]

        return get_distributions(moves, bucket_length)

    def distribution_for_up_moves(self, bucket_length: int, max_candle_size_for_analysis: float) -> dict:
        return self.distribution_for_moves('up', bucket_length, max_candle_size_for_analysis)

    def distribution_for_down_moves(self, bucket_length: int, max_candle_size_for_analysis: float) -> dict:
        return self.distribution_for_moves('down', bucket_length, max_candle_size_for_analysis)

    def distribution_for_candle_body(self, bucket_length: int,
                                     max_candle_size_for_analysis: float) -> dict:
        candle_body_lengths = self.data.get_candle_body_lengths()

        # IMPORTANT - remove too big candles, they are outliers to this analysis
        candle_body_lengths = [x for x in candle_body_lengths if x <= max_candle_size_for_analysis]

        return get_distributions(candle_body_lengths, bucket_length)

    def distribution_for_candle_lower_wick(self, bucket_length: int,
                                           max_candle_size_for_analysis: float) -> dict:
        candle_lower_wick_lengths = self.data.get_candle_lower_wick_lengths()

        # IMPORTANT - remove too big candles, they are outliers to this analysis
        candle_lower_wick_lengths = [x for x in candle_lower_wick_lengths if x <= max_candle_size_for_analysis]

        return get_distributions(candle_lower_wick_lengths, bucket_length)

    def distribution_for_candle_upper_wick(self, bucket_length: int,
                                           max_candle_size_for_analysis: float) -> dict:
        candle_upper_wick_lengths = self.data.get_candle_upper_wick_lengths()

        # IMPORTANT - remove too big candles, they are outliers to this analysis
        candle_upper_wick_lengths = [x for x in candle_upper_wick_lengths if x <= max_candle_size_for_analysis]

        return get_distributions(candle_upper_wick_lengths, bucket_length)


cache = {
    'key': {
        'unique_instrument_token': None,
        'start_date': None,
        'end_date': None,
    },
    'val': None
}


def is_key_present_in_cache(
        unique_instrument_token,
        start_date,
        end_date,
) -> bool:
    if cache['key']['unique_instrument_token'] != unique_instrument_token:
        return False
    if cache['key']['start_date'] != start_date:
        return False
    if cache['key']['end_date'] != end_date:
        return False

    return True


def fetch_candlestick_data_from_upstox(
        unique_instrument_token: str,
        start_date: date,
        end_date: date,
) -> UpstoxCandlestickResponse:
    if is_key_present_in_cache(unique_instrument_token, start_date, end_date):
        return cache['val']

    resp = requests.get(
        url='https://api.upstox.com/v2/historical-candle/{}/1minute/{}/{}'
        .format(
            unique_instrument_token,
            end_date.strftime(date_str_format),
            start_date.strftime(date_str_format),
        )
    )

    if resp.status_code != 200:
        raise Exception(f'Upstox fetch api failed, status_code: {resp.status_code}')

    resp = UpstoxCandlestickResponse.from_upstox_api_response(json.loads(resp.content))
    cache['key']['unique_instrument_token'] = unique_instrument_token
    cache['key']['start_date'] = start_date
    cache['key']['end_date'] = end_date
    cache['val'] = resp

    return resp
