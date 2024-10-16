import os
from typing import List
from openpyxl import load_workbook

from common import get_distributions, write_distribution_result_to_worksheet

input_sheet_relative_file_path = 'BANKNIFTY_historical_analysis.xlsx'
momentum_sheet_name = 'momentum_analysis'
momentum_col = 2
start_row = 3
end_row = 93
grouping_range_width = 10


def fetch_momentums_from_input_sheet() -> List[int]:
    full_path = os.getcwd() + '/' + input_sheet_relative_file_path
    wb = load_workbook(full_path)
    ws = wb[momentum_sheet_name]

    momentums = []
    for row in ws.iter_rows(min_col=momentum_col, max_col=momentum_col, min_row=start_row, max_row=end_row):
        for cell in row:
            momentums.append(cell.value)

    return momentums


def get_momentum_distributions(momentums: List[int]) -> dict:
    return get_distributions(momentums, grouping_range_width)


def write_output_to_sheet(momentum_distributions: dict):
    full_path = os.getcwd() + '/' + input_sheet_relative_file_path

    write_distribution_result_to_worksheet(
        momentum_distributions,
        full_path,
        momentum_sheet_name,
        3,
        5
    )


def main():
    momentums = fetch_momentums_from_input_sheet()
    momentum_distributions = get_momentum_distributions(momentums)
    write_output_to_sheet(momentum_distributions)


if __name__ == '__main__':
    main()
