import os
from typing import TypedDict
from datetime import date, timedelta

from config import market_instrument_token
from kiteconnect import KiteConnect

from common import new_kite_connect_client
from openpyxl import load_workbook
from typing import List


report_file_relative_path = 'NIFTY_historical_analysis.xlsx'
sheet_name = 'daily_fluctuation'
min_row = 2
max_row = 1000
min_col = 1  # A
max_col = 7  # G
from_date = date(2024, 7, 1)
to_date = date(2024, 9, 9)


# output: {date -> [low, high]}
def get_daily_candlesticks(kc: KiteConnect) -> dict:
    candle_stick = kc.historical_data(
        instrument_token=market_instrument_token,
        from_date=from_date,
        to_date=to_date,
        interval='day',
    )
    
    res = {}
    
    for candle in candle_stick:
        res[candle['date'].strftime('%Y-%m-%d')] = [candle['low'], candle['high']]

    return res


# output candle opening price at 10 am
# {date -> price}
def get_entry_prices(kc: KiteConnect) -> dict:
    res = {}

    day = from_date
    while day <= to_date:
        candle_stick = kc.historical_data(
            instrument_token=market_instrument_token,
            from_date=from_date,
            to_date=to_date,
            interval='5minute',
        )

        # pick 10 am candle
        res[day.strftime('%Y-%m-%d')] = candle_stick[3]['open']

        day += timedelta(days=1)

    return res


class Report(TypedDict):
    day: date
    entry_price: float
    high: float
    low: float
    delta_high: float
    delta_low: float
    max_delta: float


def report_to_data_list(report: Report) -> List:
    return [
        report['day'].strftime('%Y-%m-%d'),
        report['entry_price'],
        report['high'],
        report['low'],
        report['delta_high'],
        report['delta_low'],
        report['max_delta'],
    ]


def generate_reports(daily_candlesticks: dict, entry_prices: dict) -> List[Report]:
    reports = []

    day = from_date
    while day <= to_date:
        day_str = day.strftime('%Y-%m-%d')

        if day_str in daily_candlesticks:
            delta_high = abs(entry_prices[day_str] - daily_candlesticks[day_str][1])
            delta_low = abs(entry_prices[day_str] - daily_candlesticks[day_str][0])
            max_delta = max(delta_high, delta_low)

            report = Report(
                day=day,
                entry_price=entry_prices[day_str],
                high=daily_candlesticks[day_str][1],
                low=daily_candlesticks[day_str][0],
                delta_high=delta_high,
                delta_low=delta_low,
                max_delta=max_delta,
            )

            reports.append(report)

        day += timedelta(days=1)

    return reports


def save_all_reports(reports: List[Report]):
    full_path = os.getcwd() + '/' + report_file_relative_path
    wb = load_workbook(full_path)
    ws = wb[sheet_name]

    # IMPORTANT ---> in order to clear the sheet please uncomment the following code
    # for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
    #     for cell in row:
    #         cell.value = None

    for report in reports:
        ws.append(report_to_data_list(report))

    wb.save(full_path)
    wb.close()


def main():
    kc = new_kite_connect_client()

    daily_candlesticks = get_daily_candlesticks(kc)
    entry_prices = get_entry_prices(kc)

    reports: List[Report] = generate_reports(daily_candlesticks, entry_prices)
    save_all_reports(reports)


if __name__ == '__main__':
    main()
