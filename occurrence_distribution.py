import os
from typing import List
from openpyxl import load_workbook


report_file_relative_path = 'NIFTY_historical_analysis.xlsx'
daily_fluctuation_data_sheet = 'daily_fluctuation'
daily_fluctuation_distribution_curve_sheet_name = 'fluctuation_distribution'
deviation_col = 7  # G
start_row = 2
end_row = 666
grouping_range_width = 30


def fetch_daily_market_fluctuation() -> List[float]:
    full_path = os.getcwd() + '/' + report_file_relative_path
    wb = load_workbook(full_path)
    ws = wb[daily_fluctuation_data_sheet]

    daily_market_fluctuation = []
    for row in ws.iter_rows(min_col=deviation_col, max_col=deviation_col, min_row=start_row, max_row=end_row):
        for cell in row:
            daily_market_fluctuation.append(cell.value)

    return daily_market_fluctuation


# output: {"10-20": [3, cum_occurrence], "20-30": [4, cum_occurrence]}
def convert_to_occurrence_distribution(daily_market_fluctuation: List[float]) -> dict:
    daily_market_fluctuation.sort()
    occurrence_distribution = {}

    start_range = 0
    end_range = grouping_range_width

    i = 0
    sum_of_occurrence = 0
    tot_cumulative_occurrence = 0
    while i < len(daily_market_fluctuation):
        if daily_market_fluctuation[i] <= end_range:
            sum_of_occurrence += 1
        else:
            key = f'{start_range}-{end_range}'
            occurrence_distribution[key] = [sum_of_occurrence]
            tot_cumulative_occurrence += sum_of_occurrence

            start_range = end_range
            end_range = start_range + grouping_range_width

            sum_of_occurrence = 0
            i -= 1

        i += 1

    if sum_of_occurrence != 0:
        occurrence_distribution[f'{start_range}-{end_range}'] = [sum_of_occurrence]
        tot_cumulative_occurrence += sum_of_occurrence

    # calculate cumulative percentage
    start_range = 0
    end_range = grouping_range_width
    tot_sum_so_far = 0
    while True:
        my_range = f'{start_range}-{end_range}'
        if my_range not in occurrence_distribution:
            break

        count = occurrence_distribution[my_range][0]
        tot_sum_so_far += count

        cur_percentage = round(tot_sum_so_far / tot_cumulative_occurrence * 100)
        occurrence_distribution[my_range].append(cur_percentage)

        start_range = end_range
        end_range = start_range + grouping_range_width

    return occurrence_distribution


def _clean_worksheet(ws):
    row = 2
    while row <= 10000:
        for cell in ws[row]:
            cell.value = None

        row += 1


def save_to_report_sheet(occurrence_distribution: dict):
    full_path = os.getcwd() + '/' + report_file_relative_path

    wb = load_workbook(full_path)
    ws = wb[daily_fluctuation_distribution_curve_sheet_name]

    _clean_worksheet(ws)

    cur_row = 2
    for group_range, arr in occurrence_distribution.items():
        count = arr[0]
        cum_percentage = arr[1]

        ws.cell(row=cur_row, column=1, value=group_range)
        ws.cell(row=cur_row, column=2, value=count)
        ws.cell(row=cur_row, column=3, value=cum_percentage)

        cur_row += 1

    wb.save(full_path)
    wb.close()


def main():
    daily_market_fluctuation = fetch_daily_market_fluctuation()
    occurrence_distribution = convert_to_occurrence_distribution(daily_market_fluctuation)
    save_to_report_sheet(occurrence_distribution)


if __name__ == '__main__':
    main()
