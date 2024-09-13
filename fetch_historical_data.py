import os
from typing import TypedDict
from datetime import date, timedelta

from config import market_instrument_token
from kiteconnect import KiteConnect

from common import new_kite_connect_client
from openpyxl import load_workbook
from typing import List


report_file_relative_path = 'NIFTY_historical_analysis.xlsx'
sheet_name = 'daily_fluctuation'
min_row = 2
max_row = 10000
min_col = 1  # A
max_col = 7  # G


# output: {date -> [low, high]}
def get_daily_candlesticks(kc: KiteConnect, from_date: date, to_date: date) -> dict:
    candle_stick = kc.historical_data(
        instrument_token=market_instrument_token,
        from_date=from_date,
        to_date=to_date,
        interval='day',
    )
    
    res = {}
    
    for candle in candle_stick:
        res[candle['date'].strftime('%Y-%m-%d')] = [candle['low'], candle['high']]

    return res


# output candle opening price at 10 am
# {date -> price}
def get_entry_prices(kc: KiteConnect, from_date: date, to_date: date) -> dict:
    res = {}

    day = from_date
    while day <= to_date:
        candle_stick = kc.historical_data(
            instrument_token=market_instrument_token,
            from_date=day,
            to_date=day,
            interval='15minute',
        )

        # pick 10 am candle
        if len(candle_stick) > 0:  # If not a trade holiday
            res[day.strftime('%Y-%m-%d')] = candle_stick[3]['open']

        day += timedelta(days=1)

    return res


class Report(TypedDict):
    day: date
    entry_price: float
    high: float
    low: float
    delta_high: float
    delta_low: float
    max_delta: float


def report_to_data_list(report: Report) -> List:
    return [
        report['day'].strftime('%Y-%m-%d'),
        report['entry_price'],
        report['high'],
        report['low'],
        report['delta_high'],
        report['delta_low'],
        report['max_delta'],
    ]


def generate_reports(daily_candlesticks: dict, entry_prices: dict, from_date: date, to_date: date) -> List[Report]:
    reports = []

    day = from_date
    while day <= to_date:
        day_str = day.strftime('%Y-%m-%d')

        if day_str in daily_candlesticks:
            delta_high = abs(entry_prices[day_str] - daily_candlesticks[day_str][1])
            delta_low = abs(entry_prices[day_str] - daily_candlesticks[day_str][0])
            max_delta = max(delta_high, delta_low)

            report = Report(
                day=day,
                entry_price=entry_prices[day_str],
                high=daily_candlesticks[day_str][1],
                low=daily_candlesticks[day_str][0],
                delta_high=delta_high,
                delta_low=delta_low,
                max_delta=max_delta,
            )

            reports.append(report)

        day += timedelta(days=1)

    return reports


def clean_worksheet():
    full_path = os.getcwd() + '/' + report_file_relative_path
    wb = load_workbook(full_path)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None

    wb.save(full_path)
    wb.close()


def save_all_reports(reports: List[Report], cur_row: int):
    full_path = os.getcwd() + '/' + report_file_relative_path
    wb = load_workbook(full_path)
    ws = wb[sheet_name]

    for report in reports:
        data_list = report_to_data_list(report)

        ws.cell(row=cur_row, column=1, value=data_list[0])
        ws.cell(row=cur_row, column=2, value=data_list[1])
        ws.cell(row=cur_row, column=3, value=data_list[2])
        ws.cell(row=cur_row, column=4, value=data_list[3])
        ws.cell(row=cur_row, column=5, value=data_list[4])
        ws.cell(row=cur_row, column=6, value=data_list[5])
        ws.cell(row=cur_row, column=7, value=data_list[6])

        cur_row += 1

    wb.save(full_path)
    wb.close()

    return cur_row


def generate_for_specific_date_range(kc: KiteConnect, from_date: date, to_date: date, start_row: int) -> int:
    daily_candlesticks = get_daily_candlesticks(kc, from_date, to_date)
    entry_prices = get_entry_prices(kc, from_date, to_date)

    reports: List[Report] = generate_reports(daily_candlesticks, entry_prices, from_date, to_date)
    next_start_row = save_all_reports(reports, start_row)

    return next_start_row


def main():
    date_ranges = [
        [date(2022, 1, 1), date(2022, 3, 30)],
        [date(2022, 4, 1), date(2022, 6, 30)],
        [date(2022, 7, 1), date(2022, 9, 30)],
        [date(2022, 10, 1), date(2022, 12, 30)],
        [date(2023, 1, 1), date(2023, 3, 30)],
        [date(2023, 4, 1), date(2023, 6, 30)],
        [date(2023, 7, 1), date(2023, 9, 30)],
        [date(2023, 10, 1), date(2023, 12, 30)],
        [date(2024, 1, 1), date(2024, 3, 30)],
        [date(2024, 4, 1), date(2024, 6, 30)],
        [date(2024, 7, 1), date(2024, 9, 9)],
    ]

    clean_worksheet()

    kc = new_kite_connect_client()

    start_row = 2
    for date_range in date_ranges:
        next_row = generate_for_specific_date_range(kc, date_range[0], date_range[1], start_row)
        start_row = next_row


if __name__ == '__main__':
    main()
