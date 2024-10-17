import math
import os
from datetime import datetime, timedelta, date, time
from typing import List
from openpyxl import load_workbook
from common import fetch_candlestick_data_from_upstox, UpstoxCandlestickResponse, Candle, time_str_format, \
    date_str_format

# User input -----------------------------
NIFTY_INSTRUMENT_TOKEN = "NSE_INDEX%7CNifty%2050"

end_datetime = datetime.today()
start_datetime = end_datetime - timedelta(days=180)
unique_instrument_token = NIFTY_INSTRUMENT_TOKEN
sheet_rel_file_path = 'nifty_gap_opening_trading_analysis_report.xlsx'
trading_report_sheet_name = 'trading_report'
cell_winning_trade_day_count = 'N3'
cell_loosing_trade_day_count = 'N4'
cell_no_trade_day_count = 'N5'
cell_wining_rate = 'N6'
cell_avg_loosing_streaks = 'N7'
cell_max_loosing_streaks = 'N8'
cell_tot_gain = 'N9'
cell_gap_threshold = 'K3'
cell_fixed_sl_points = 'K4'
cell_target_type = 'K5'
cell_fixed_target_points = 'K6'


class TradeConfig:
    def __init__(
            self,
            gap_threshold: int,
            fixed_sl_points: int,
            target_type: str,
            fixed_target_points: int,
    ):
        self.gap_threshold = gap_threshold
        self.fixed_sl_points = fixed_sl_points
        self.target_type = target_type
        self.fixed_target_points = fixed_target_points
# ------------------------------------------
# ------------------------------------------


# -------------- CODE ------------------------------------------------
# --------------------------------------------------------------------
start_date: date = start_datetime.date()
end_date: date = end_datetime.date()


class Trade:
    def __init__(
            self,
            day: date,
            gap: float,
            entry_time: time = None,
            exit_time: time = None,
            trading_status: str = 'No trading',
            points_gained: float = ''
    ):
        self.day = day
        self.gap = gap
        self.entry_time = entry_time
        self.exit_time = exit_time
        self.trading_status = trading_status
        self.points_gained = points_gained


class TradingSummary:
    def __init__(
            self,
            no_trade_day_count: int,
            winning_trade_day_count: int,
            loosing_trade_day_count: int,
            wining_rate: float,
            avg_loosing_streaks: int,
            max_loosing_streaks: int,
            tot_gain: int,
    ):
        self.no_trade_day_count = no_trade_day_count
        self.winning_trade_day_count = winning_trade_day_count
        self.loosing_trade_day_count = loosing_trade_day_count
        self.wining_rate = wining_rate
        self.avg_loosing_streaks = avg_loosing_streaks
        self.max_loosing_streaks = max_loosing_streaks
        self.tot_gain = tot_gain

    @classmethod
    def from_trades(cls, trades: List[Trade]):
        no_trade_day_count = 0
        winning_trade_day_count = 0
        loosing_trade_day_count = 0

        tot_gain = 0

        n = len(trades)
        i = 0
        cur_loosing_streaks = 0
        loosing_streaks: List[int] = []
        while i < n:
            trade = trades[i]
            tot_gain += trade.points_gained if trade.points_gained != '' else 0

            if trade.trading_status == 'No trading':
                no_trade_day_count += 1
                cur_loosing_streaks = 0
            elif trade.points_gained > 0:
                winning_trade_day_count += 1
                cur_loosing_streaks = 0
            elif trade.points_gained <= 0:
                loosing_trade_day_count += 1

                # handle loosing streak related things
                cur_loosing_streaks += 1
                if len(loosing_streaks) == 0:
                    loosing_streaks.append(cur_loosing_streaks)
                elif cur_loosing_streaks == 1:
                    loosing_streaks.append(cur_loosing_streaks)
                else:
                    loosing_streaks.pop()
                    loosing_streaks.append(cur_loosing_streaks)

            i += 1

        try:
            wining_rate = round((winning_trade_day_count / (winning_trade_day_count+loosing_trade_day_count)) * 100, 2)
        except Exception:
            wining_rate = 'Nan'

        try:
            avg_loosing_streaks_cnt = math.ceil(sum(loosing_streaks) / len(loosing_streaks))
        except Exception:
            avg_loosing_streaks_cnt = 'Nan'

        try:
            max_loosing_streaks_cnt = max(loosing_streaks)
        except Exception:
            max_loosing_streaks_cnt = 'Nan'

        return TradingSummary(
            no_trade_day_count,
            winning_trade_day_count,
            loosing_trade_day_count,
            wining_rate,
            avg_loosing_streaks_cnt,
            max_loosing_streaks_cnt,
            tot_gain,
        )


def clean_worksheet(ws):
    # 1. trades
    for row in range(2, 1000):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.value = None

    # 2. trade configs
    ws[cell_gap_threshold] = None
    ws[cell_fixed_sl_points] = None
    ws[cell_target_type] = None
    ws[cell_fixed_target_points] = None

    # 3. trading summary
    ws[cell_winning_trade_day_count] = None
    ws[cell_loosing_trade_day_count] = None
    ws[cell_no_trade_day_count] = None
    ws[cell_wining_rate] = None
    ws[cell_avg_loosing_streaks] = None
    ws[cell_max_loosing_streaks] = None
    ws[cell_tot_gain] = None


def write_to_sheet(
        trades: List[Trade],
        trade_config: TradeConfig,
        trade_summary: TradingSummary,
):
    full_path = os.getcwd() + '/' + sheet_rel_file_path

    wb = load_workbook(full_path)
    try:
        ws = wb[trading_report_sheet_name]
        clean_worksheet(ws)

        # 1. trades
        cur_row = 2
        for trade in trades:
            ws.cell(row=cur_row, column=1, value=trade.day.strftime(date_str_format))
            ws.cell(row=cur_row, column=2, value=trade.gap)
            ws.cell(row=cur_row, column=3, value=trade.trading_status)

            if trade.entry_time is not None:
                ws.cell(row=cur_row, column=4, value=trade.entry_time.strftime(time_str_format))
            if trade.exit_time is not None:
                ws.cell(row=cur_row, column=5, value=trade.exit_time.strftime(time_str_format))

            ws.cell(row=cur_row, column=6, value=trade.points_gained)

            cur_row += 1

        # 2. trade configs
        ws[cell_gap_threshold] = trade_config.gap_threshold
        ws[cell_fixed_sl_points] = trade_config.fixed_sl_points
        ws[cell_target_type] = trade_config.target_type
        ws[cell_fixed_target_points] = trade_config.fixed_target_points

        # 3. trading summary
        ws[cell_winning_trade_day_count] = trade_summary.winning_trade_day_count
        ws[cell_loosing_trade_day_count] = trade_summary.loosing_trade_day_count
        ws[cell_no_trade_day_count] = trade_summary.no_trade_day_count
        ws[cell_wining_rate] = trade_summary.wining_rate
        ws[cell_avg_loosing_streaks] = trade_summary.avg_loosing_streaks
        ws[cell_max_loosing_streaks] = trade_summary.max_loosing_streaks
        ws[cell_tot_gain] = trade_summary.tot_gain
    finally:
        wb.save(full_path)
        wb.close()


def get_trades_and_summary(
        gap_threshold: int,
        fixed_sl_points: int,
        fixed_target_points: int,
) -> (List[Trade], TradingSummary):
    candlestick_data: UpstoxCandlestickResponse = fetch_candlestick_data_from_upstox(
        unique_instrument_token,
        start_date,
        end_date,
    )

    candles: List[Candle] = candlestick_data.data.candles
    n = len(candles)

    trades: List[Trade] = []

    # ------------- simulate the trades -------------
    i = 1
    while i < n-1:
        cur_candle = candles[i]
        prev_candle = candles[i-1]
        next_candle = candles[i+1]

        if cur_candle.date != prev_candle.date:
            # taking 9:16 am candle for considering gap opening not 9:15
            gap = next_candle.open - prev_candle.close
            trade = Trade(
                day=cur_candle.ts.date(),
                gap=gap,
            )

            # entry at 9:16 am
            entry_point = next_candle.open
            j = i+1

            # entry at 9:15 am
            # entry_point = cur_candle.open
            # j = i

            if gap >= abs(gap_threshold):
                # take short entry
                trade.entry_time = candles[j].ts.time()
                while j < n and candles[j].date == cur_candle.date:
                    if candles[j].hi >= entry_point + fixed_sl_points:
                        # SL hit
                        trade.exit_time = candles[j].ts.time()
                        trade.points_gained = -1 * fixed_sl_points
                        trade.trading_status = ''
                        break
                    elif candles[j].lo <= entry_point - fixed_target_points:
                        # Target hit
                        trade.exit_time = candles[j].ts.time()
                        trade.points_gained = fixed_target_points
                        trade.trading_status = ''
                        break
                    else:
                        j += 1
            elif gap <= -1 * abs(gap_threshold):
                # take long entry
                trade.entry_time = candles[j].ts.time()
                while j < n and candles[j].date == cur_candle.date:
                    if candles[j].lo <= entry_point - fixed_sl_points:
                        # SL hit
                        trade.exit_time = candles[j].ts.time()
                        trade.points_gained = -1 * fixed_sl_points
                        trade.trading_status = ''
                        break
                    elif candles[j].lo >= entry_point + fixed_target_points:
                        # Target hit
                        trade.exit_time = candles[j].ts.time()
                        trade.points_gained = fixed_target_points
                        trade.trading_status = ''
                        break
                    else:
                        j += 1

            trades.append(trade)

        i += 1

    return trades, TradingSummary.from_trades(trades)


def main(trade_config: TradeConfig):
    trades, trading_summary = get_trades_and_summary(
        trade_config.gap_threshold,
        trade_config.fixed_sl_points,
        trade_config.fixed_target_points,
    )

    write_to_sheet(trades, trade_config, trading_summary)


def optimization() -> TradeConfig:
    gap_threshold_range = list(range(30, 201, 10))
    fixed_sl_points_range = list(range(20, 41, 5))
    fixed_target_points_range = list(range(30, 81, 5))

    max_tot_gain = -10000000
    optimized_trade_config = None

    for gap_threshold in gap_threshold_range:
        for fixed_sl_points in fixed_sl_points_range:
            for fixed_target_points in fixed_target_points_range:
                trades, trading_summary = get_trades_and_summary(
                    gap_threshold,
                    fixed_sl_points,
                    fixed_target_points,
                )

                if max_tot_gain < trading_summary.tot_gain:
                    max_tot_gain = trading_summary.tot_gain
                    optimized_trade_config = TradeConfig(
                        gap_threshold=gap_threshold,
                        fixed_sl_points=fixed_sl_points,
                        target_type='fixed',
                        fixed_target_points=fixed_target_points,
                    )

    return optimized_trade_config


if __name__ == '__main__':
    # manual optimisation and testing
    # main(TradeConfig(
    #     gap_threshold=80,
    #     fixed_sl_points=40,
    #     target_type='fixed',
    #     fixed_target_points=80,
    # ))

    # most optimised one as per "tot_gain"
    main(TradeConfig(
        gap_threshold=70,
        fixed_sl_points=40,
        target_type='fixed',
        fixed_target_points=80,
    ))

    optimised_trade_config: TradeConfig = optimization()
    print(optimised_trade_config.__dict__)
